from openpyxl import load_workbook

from src.db.database import get_connection


def save_lotto_number(lotto_data: dict):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT OR REPLACE INTO lotto_winning_numbers (
            draw_no,
            draw_date,
            number1,
            number2,
            number3,
            number4,
            number5,
            number6,
            bonus_number
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        lotto_data["draw_no"],
        lotto_data["draw_date"],
        lotto_data["number1"],
        lotto_data["number2"],
        lotto_data["number3"],
        lotto_data["number4"],
        lotto_data["number5"],
        lotto_data["number6"],
        lotto_data["bonus_number"],
    ))

    conn.commit()
    conn.close()


def collect_from_excel(file_path: str):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    saved_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            continue

        lotto_data = {
            "draw_no": int(row[1]),
            "draw_date": "",
            "number1": int(row[2]),
            "number2": int(row[3]),
            "number3": int(row[4]),
            "number4": int(row[5]),
            "number5": int(row[6]),
            "number6": int(row[7]),
            "bonus_number": int(row[8]),
        }

        save_lotto_number(lotto_data)
        saved_count += 1

    print(f"엑셀 당첨번호 저장 완료: {saved_count}건")